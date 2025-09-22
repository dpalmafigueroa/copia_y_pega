import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import base64

st.set_page_config(layout="wide")

st.title("Pegado de datos automatizado dentro de un mismo archivo Excel 📂")
st.write("Sube tu archivo para pegar datos de una hoja a otra.")

# --- LÓGICA DE PROCESAMIENTO ---
@st.cache_data
def obtener_nombres_de_hojas(uploaded_file):
    """Función para obtener los nombres de todas las hojas de un archivo de Excel."""
    try:
        xls = pd.ExcelFile(uploaded_file)
        return xls.sheet_names
    except Exception as e:
        st.error(f"❌ No se pudo leer el archivo: {e}")
        return []

def procesar_hoja(single_file, source_sheet, target_sheet, headers_row, start_row):
    try:
        # Lee los datos de la hoja de origen con pandas
        df_base = pd.read_excel(single_file, sheet_name=source_sheet, engine="openpyxl")
        
        # Carga el libro de trabajo con openpyxl para modificar la hoja de destino
        wb = load_workbook(single_file)
        ws = wb[target_sheet]

        # Obtiene los encabezados de la hoja de destino
        headers_plantilla = {
            str(cell.value).strip(): cell.column
            for cell in ws[int(headers_row)] if cell.value
        }
        
        # Encuentra las columnas que coinciden
        columnas_comunes = [col for col in df_base.columns if col in headers_plantilla]
        
        if not columnas_comunes:
            raise ValueError("No se encontraron columnas coincidentes entre las hojas de origen y destino.")

        df_filtrado = df_base[columnas_comunes]

        # Prepara los datos para ser pegados
        rows_to_paste = dataframe_to_rows(df_filtrado, index=False, header=False)
        
        # Pega los datos en la hoja de destino
        for r_idx, row in enumerate(rows_to_paste, start=int(start_row)):
            for col_name, value in zip(df_filtrado.columns, row):
                col_idx = headers_plantilla[col_name]
                ws.cell(row=r_idx, column=col_idx, value=value)

        # Guarda el archivo modificado en un buffer de memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, len(df_filtrado)

    except KeyError as e:
        st.error(f"❌ Error: La hoja '{e.args[0]}' no existe en el archivo. Por favor, verifica el nombre.")
        return None, 0
    except Exception as e:
        st.error(f"❌ Ocurrió un error inesperado durante el procesamiento: {e}")
        return None, 0

# --- INTERFAZ DE USUARIO ---
with st.form(key='my_form'):
    uploaded_file = st.file_uploader("Sube tu archivo Excel", type=["xlsx"])
    
    source_sheet_name = None
    target_sheet_name = None

    if uploaded_file:
        sheet_names = obtener_nombres_de_hojas(uploaded_file)
        if sheet_names:
            source_sheet_name = st.selectbox(
                "Selecciona la hoja de origen (con los datos a copiar):",
                sheet_names
            )
            target_sheet_name = st.selectbox(
                "Selecciona la hoja de destino (donde se pegarán los datos):",
                sheet_names
            )
        else:
            st.warning("No se pudieron obtener las hojas del archivo.")

    headers_row = st.number_input("Ingresa la fila de encabezados en la hoja de destino:", min_value=1, value=1)
    start_row = st.number_input("Ingresa la fila de inicio para el pegado:", min_value=1, value=2)
    
    submit_button = st.form_submit_button("🚀 Procesar y Pegar Datos")

if submit_button:
    if uploaded_file and source_sheet_name and target_sheet_name:
        if source_sheet_name == target_sheet_name:
            st.error("❌ La hoja de origen y la hoja de destino no pueden ser la misma.")
        else:
            with st.spinner("Procesando... por favor, espera."):
                output_file, longitud_max = procesar_hoja(
                    uploaded_file,
                    source_sheet_name,
                    target_sheet_name,
                    headers_row,
                    start_row
                )
            
            if output_file:
                st.success(f"✅ ¡Pegado de {longitud_max} filas completado desde la fila {start_row}!")
                b64 = base64.b64encode(output_file.getvalue()).decode()
                href = f'<a href="data:application/octet-stream;base64,{b64}" download="archivo_modificado.xlsx">📥 Descargar archivo modificado</a>'
                st.markdown(href, unsafe_allow_html=True)
    else:
        st.error("❌ Por favor, sube el archivo y selecciona las hojas.")
